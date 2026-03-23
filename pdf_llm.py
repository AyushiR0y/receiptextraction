#!/usr/bin/env python3
"""
Receipt Data Extractor
Processes PDF, JPG, and ZIP files containing insurance/brokerage receipts
and extracts structured data into an Excel spreadsheet.

Usage:
    python receipt_extractor.py <input_folder> [output.xlsx]
    python receipt_extractor.py invoice.pdf [output.xlsx]
    python receipt_extractor.py receipts.zip [output.xlsx]
"""

import os
import sys
import re
import json
import zipfile
import tempfile
import argparse
import base64
from pathlib import Path
from datetime import datetime

import pdfplumber
from pypdf import PdfReader, PdfWriter
from PIL import Image
import pytesseract
from pdf2image import convert_from_path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── Password map for protected PDFs ────────────────────────────────────────
BANK_PASSWORDS = {
    "axis bank":   "AADCA1701E",
    "dbs":         "db$126497",
    "marsh india": "__Marsh@2025__",
}

# ─── Excel column headers ────────────────────────────────────────────────────
COLUMNS = [
    "AGENT_CODE", "Agent Name", "Agent PAN", "Name of Service Receipient",
    "BALIC STATE", "BALIC GSTN", "BROKER GSTN STATE", "BROKER GSTN",
    "Vendor Inv Date", "Vendor Inv No", "Total Inv Amt", "BROKERAGE Amount",
    "CGST @ 9%", "SGST @ 9%", "UTGST", "IGST", "GST TOTAL AMT",
    "DATE_FROM", "DATE_TO", "Narration", "Type", "Micro/Non Micro", "SAC Code",
]

# ─── Field extraction patterns (regex fallback) ─────────────────────────────
PATTERNS = {
    "Vendor Inv No":           r"(?:invoice\s*(?:no|number|#)[:\s]*)([\w\-/]+)",
    "Vendor Inv Date":         r"(?:invoice\s*date|date\s*of\s*invoice)[:\s]*([\d]{1,2}[-/\s][\w]{2,9}[-/\s][\d]{2,4}|[\d]{1,2}[-/][\d]{1,2}[-/][\d]{2,4})",
    "Total Inv Amt":           r"(?:total\s*(?:invoice\s*)?(?:amount|amt|value)|grand\s*total)[:\s]*(?:rs\.?|inr|₹)?\s*([\d,]+(?:\.\d{1,2})?)",
    "BROKERAGE Amount":        r"(?:brokerage\s*(?:amount|amt)?)[:\s]*(?:rs\.?|inr|₹)?\s*([\d,]+(?:\.\d{1,2})?)",
    "CGST @ 9%":               r"cgst\s*(?:@\s*9%?)?[:\s]*(?:rs\.?|inr|₹)?\s*([\d,]+(?:\.\d{1,2})?)",
    "SGST @ 9%":               r"sgst\s*(?:@\s*9%?)?[:\s]*(?:rs\.?|inr|₹)?\s*([\d,]+(?:\.\d{1,2})?)",
    "UTGST":                   r"utgst[:\s]*(?:rs\.?|inr|₹)?\s*([\d,]+(?:\.\d{1,2})?)",
    "IGST":                    r"igst[:\s]*(?:rs\.?|inr|₹)?\s*([\d,]+(?:\.\d{1,2})?)",
    "GST TOTAL AMT":           r"(?:total\s*gst|gst\s*(?:total|amount|amt))[:\s]*(?:rs\.?|inr|₹)?\s*([\d,]+(?:\.\d{1,2})?)",
    "Agent PAN":               r"\b([A-Z]{5}[0-9]{4}[A-Z])\b",
    "BROKER GSTN":             r"\b(\d{2}[A-Z]{5}\d{4}[A-Z]\d[Z][A-Z\d])\b",
    "BALIC GSTN":              r"(?:balic|insurer|company)\s*gstn?[:\s]*([\d]{2}[A-Z]{5}[\d]{4}[A-Z][\dA-Z]{3})",
    "SAC Code":                r"(?:sac\s*(?:code)?|service\s*accounting\s*code)[:\s]*(\d{4,8})",
    "DATE_FROM":               r"(?:from\s*date|date\s*from|period\s*from|coverage\s*from)[:\s]*([\d]{1,2}[-/][\d]{1,2}[-/][\d]{2,4}|[\d]{1,2}[-/\s][\w]+[-/\s][\d]{2,4})",
    "DATE_TO":                 r"(?:to\s*date|date\s*to|period\s*to|coverage\s*to)[:\s]*([\d]{1,2}[-/][\d]{1,2}[-/][\d]{2,4}|[\d]{1,2}[-/\s][\w]+[-/\s][\d]{2,4})",
    "Agent Name":              r"(?:agent\s*name|broker\s*name)[:\s]*([A-Za-z\s\.]+?)(?:\n|pan|gstin|$)",
    "AGENT_CODE":              r"(?:agent\s*code|broker\s*code)[:\s]*([\w\-]+)",
    "Name of Service Receipient": r"(?:insured\s*name|name\s*of\s*insured|policyholder|service\s*recipient)[:\s]*([A-Za-z\s\.]+?)(?:\n|pan|gstin|$)",
    "BALIC STATE":             r"(?:balic\s*state|insurer\s*state)[:\s]*([A-Za-z\s]+?)(?:\n|gstin|$)",
    "BROKER GSTN STATE":       r"(?:broker\s*state|broker\s*gstn\s*state)[:\s]*([A-Za-z\s]+?)(?:\n|gstin|$)",
    "Narration":               r"(?:narration|description\s*of\s*service)[:\s]*(.+?)(?:\n\n|$)",
    "Type":                    r"(?:type\s*of\s*(?:policy|insurance|service))[:\s]*([A-Za-z\s]+?)(?:\n|$)",
    "Micro/Non Micro":         r"(micro|non[\s\-]?micro)",
}


# ─── Utility ─────────────────────────────────────────────────────────────────

def try_decrypt_pdf(pdf_path: str) -> tuple[PdfReader | None, str]:
    """Try to open a PDF, decrypting with known passwords if needed."""
    try:
        reader = PdfReader(pdf_path)
        if reader.is_encrypted:
            # Try each known password
            for bank, pwd in BANK_PASSWORDS.items():
                try:
                    result = reader.decrypt(pwd)
                    if result:
                        return reader, bank
                except Exception:
                    continue
            # Try empty password
            try:
                reader.decrypt("")
                return reader, "unknown"
            except Exception:
                pass
            return None, "locked"
        return reader, ""
    except Exception as e:
        return None, str(e)


def extract_text_from_pdf_page(page) -> str:
    """Extract text from a pdfplumber page, trying OCR if text is sparse."""
    text = page.extract_text() or ""
    if len(text.strip()) < 50:
        # Sparse text → render page to image and OCR
        img = page.to_image(resolution=200)
        pil_img = img.original
        text = pytesseract.image_to_string(pil_img, lang="eng")
    return text


def extract_text_from_image(image_path: str) -> str:
    """OCR an image file."""
    img = Image.open(image_path)
    # Auto-rotate if needed
    img = img.convert("RGB")
    text = pytesseract.image_to_string(img, lang="eng")
    return text


def clean_amount(val: str) -> str:
    """Remove commas/spaces from monetary values."""
    if val:
        return val.replace(",", "").strip()
    return val


def parse_with_regex(text: str) -> dict:
    """Run all regex patterns against extracted text and return matches."""
    text_lower = text.lower()
    result = {}
    for field, pattern in PATTERNS.items():
        m = re.search(pattern, text_lower, re.IGNORECASE | re.MULTILINE)
        if m:
            val = m.group(1).strip()
            # Restore original case for PAN/GSTN
            if field in ("Agent PAN", "BROKER GSTN", "BALIC GSTN"):
                # re-search in original text for casing
                m2 = re.search(pattern, text, re.IGNORECASE | re.MULTILINE)
                val = m2.group(1).strip() if m2 else val.upper()
            if field in ("Total Inv Amt", "BROKERAGE Amount", "CGST @ 9%",
                         "SGST @ 9%", "UTGST", "IGST", "GST TOTAL AMT"):
                val = clean_amount(val)
            result[field] = val
    return result


def call_claude_api(text_or_image: str, is_image: bool = False) -> dict:
    """
    Call Claude Sonnet via Anthropic API for structured extraction.
    Returns a dict of field → value.
    NOTE: This function is intended to be called from an environment
    where the Anthropic SDK is available. In CLI mode we fall back to regex.
    """
    try:
        import anthropic
        client = anthropic.Anthropic()

        system_prompt = """You are an expert at extracting structured data from Indian insurance brokerage receipts/invoices.
Extract ALL of the following fields from the document text. Return ONLY a valid JSON object with these exact keys (use null for missing fields):

AGENT_CODE, Agent Name, Agent PAN, Name of Service Receipient, BALIC STATE, BALIC GSTN,
BROKER GSTN STATE, BROKER GSTN, Vendor Inv Date, Vendor Inv No, Total Inv Amt,
BROKERAGE Amount, CGST @ 9%, SGST @ 9%, UTGST, IGST, GST TOTAL AMT,
DATE_FROM, DATE_TO, Narration, Type, Micro/Non Micro, SAC Code

Rules:
- Monetary amounts: numbers only (no ₹ or commas), e.g. "12500.00"
- Dates: DD/MM/YYYY format
- PAN: 10-char alphanumeric e.g. ABCDE1234F
- GSTN: 15-char alphanumeric
- Return ONLY the JSON object, no markdown, no extra text"""

        if is_image:
            messages = [{"role": "user", "content": [
                {"type": "image", "source": {"type": "base64", "media_type": "image/png",
                                              "data": text_or_image}},
                {"type": "text", "text": "Extract all receipt/invoice fields from this image."}
            ]}]
        else:
            messages = [{"role": "user", "content": f"Extract all receipt fields from this text:\n\n{text_or_image[:8000]}"}]

        response = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=1000,
            system=system_prompt,
            messages=messages,
        )
        raw = response.content[0].text.strip()
        # Strip markdown fences if any
        raw = re.sub(r"```json|```", "", raw).strip()
        data = json.loads(raw)
        # Clean up nulls
        return {k: (str(v) if v is not None else "") for k, v in data.items()}
    except ImportError:
        return {}
    except Exception as e:
        print(f"  [API warning] {e}", file=sys.stderr)
        return {}


def extract_from_text(text: str, source_label: str = "") -> dict:
    """Extract fields: try Claude API first, fallback to regex."""
    if not text.strip():
        return {col: "" for col in COLUMNS}

    # Try Claude
    result = call_claude_api(text)

    # If Claude unavailable or returned nothing, use regex
    if not result:
        print(f"  Using regex extraction for {source_label}")
        result = parse_with_regex(text)

    # Ensure all columns present
    row = {}
    for col in COLUMNS:
        row[col] = result.get(col, "") or ""
    return row


# ─── File processors ─────────────────────────────────────────────────────────

def process_pdf(pdf_path: str) -> list[dict]:
    """Process a PDF file, returning one dict per page/receipt."""
    rows = []
    reader, bank_key = try_decrypt_pdf(pdf_path)

    if reader is None:
        print(f"  [ERROR] Cannot decrypt {Path(pdf_path).name} (tried all passwords)")
        return []

    if bank_key:
        print(f"  Decrypted with password for: {bank_key}")

    # Write decrypted PDF to temp file for pdfplumber
    if bank_key:
        tmp = tempfile.NamedTemporaryFile(suffix=".pdf", delete=False)
        writer = PdfWriter()
        for page in reader.pages:
            writer.add_page(page)
        writer.write(tmp.name)
        tmp.close()
        work_path = tmp.name
    else:
        work_path = pdf_path

    try:
        with pdfplumber.open(work_path) as pdf:
            total_pages = len(pdf.pages)
            print(f"  {total_pages} page(s) found in {Path(pdf_path).name}")
            for i, page in enumerate(pdf.pages, 1):
                print(f"  Processing page {i}/{total_pages}...")
                text = extract_text_from_pdf_page(page)
                row = extract_from_text(text, f"{Path(pdf_path).name} p{i}")
                row["_source"] = f"{Path(pdf_path).name} | Page {i}"
                rows.append(row)
    finally:
        if bank_key and os.path.exists(work_path):
            os.unlink(work_path)

    return rows


def process_image(image_path: str) -> list[dict]:
    """Process a JPG/PNG image file."""
    print(f"  OCR-ing image: {Path(image_path).name}")
    text = extract_text_from_image(image_path)
    row = extract_from_text(text, Path(image_path).name)
    row["_source"] = Path(image_path).name
    return [row]


def process_zip(zip_path: str) -> list[dict]:
    """Unpack a ZIP and process all PDFs and images inside."""
    rows = []
    with tempfile.TemporaryDirectory() as tmpdir:
        with zipfile.ZipFile(zip_path, "r") as zf:
            zf.extractall(tmpdir)

        for root, _, files in os.walk(tmpdir):
            for fname in sorted(files):
                fpath = os.path.join(root, fname)
                ext = Path(fname).suffix.lower()
                print(f"  [ZIP] Processing: {fname}")
                if ext == ".pdf":
                    rows.extend(process_pdf(fpath))
                elif ext in (".jpg", ".jpeg", ".png", ".tiff", ".bmp"):
                    rows.extend(process_image(fpath))
                else:
                    print(f"  [ZIP] Skipping unsupported file: {fname}")
    return rows


def process_path(input_path: str) -> list[dict]:
    """Dispatch to the right processor based on file type or folder."""
    path = Path(input_path)
    if path.is_dir():
        rows = []
        for f in sorted(path.iterdir()):
            print(f"\nProcessing: {f.name}")
            rows.extend(process_path(str(f)))
        return rows
    ext = path.suffix.lower()
    if ext == ".pdf":
        return process_pdf(input_path)
    elif ext in (".jpg", ".jpeg", ".png", ".tiff", ".bmp"):
        return process_image(input_path)
    elif ext == ".zip":
        return process_zip(input_path)
    else:
        print(f"  [SKIP] Unsupported file type: {path.name}")
        return []


# ─── Excel writer ─────────────────────────────────────────────────────────────

def write_excel(rows: list[dict], output_path: str):
    """Write extracted rows to a formatted Excel file."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Receipts"

    # Header styling
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill("solid", start_color="1F4E79")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    display_cols = COLUMNS + ["_source"]
    header_labels = COLUMNS + ["Source File"]

    # Write headers
    for col_idx, label in enumerate(header_labels, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    ws.row_dimensions[1].height = 30

    # Alternating row fills
    fill_light = PatternFill("solid", start_color="EBF3FB")
    fill_white = PatternFill("solid", start_color="FFFFFF")
    data_font = Font(name="Arial", size=9)
    data_align = Alignment(vertical="center", wrap_text=False)

    for row_idx, row_data in enumerate(rows, 2):
        fill = fill_light if row_idx % 2 == 0 else fill_white
        for col_idx, col_key in enumerate(display_cols, 1):
            val = row_data.get(col_key, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = data_font
            cell.fill = fill
            cell.alignment = data_align
            cell.border = border

    # Auto-fit column widths
    col_widths = {
        "AGENT_CODE": 14, "Agent Name": 22, "Agent PAN": 14,
        "Name of Service Receipient": 28, "BALIC STATE": 16, "BALIC GSTN": 20,
        "BROKER GSTN STATE": 18, "BROKER GSTN": 20, "Vendor Inv Date": 16,
        "Vendor Inv No": 18, "Total Inv Amt": 15, "BROKERAGE Amount": 18,
        "CGST @ 9%": 12, "SGST @ 9%": 12, "UTGST": 12, "IGST": 12,
        "GST TOTAL AMT": 15, "DATE_FROM": 14, "DATE_TO": 14,
        "Narration": 30, "Type": 16, "Micro/Non Micro": 16, "SAC Code": 12,
        "Source File": 35,
    }
    for col_idx, label in enumerate(header_labels, 1):
        key = COLUMNS[col_idx - 1] if col_idx <= len(COLUMNS) else "Source File"
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(key, 15)

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = ws.dimensions

    wb.save(output_path)
    print(f"\n✅ Excel saved to: {output_path}")
    print(f"   Total rows extracted: {len(rows)}")


# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Extract receipt data from PDF/JPG/ZIP files into Excel"
    )
    parser.add_argument("input", help="Input file (PDF/JPG/ZIP) or folder containing files")
    parser.add_argument("output", nargs="?", default="receipts_output.xlsx",
                        help="Output Excel file (default: receipts_output.xlsx)")
    args = parser.parse_args()

    if not os.path.exists(args.input):
        print(f"Error: Input path not found: {args.input}", file=sys.stderr)
        sys.exit(1)

    print(f"Receipt Extractor")
    print(f"=================")
    print(f"Input : {args.input}")
    print(f"Output: {args.output}")
    print()

    rows = process_path(args.input)

    if not rows:
        print("\n⚠️  No data extracted. Check that your files are readable.")
        sys.exit(1)

    write_excel(rows, args.output)


if __name__ == "__main__":
    main()